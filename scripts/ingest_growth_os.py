#!/usr/bin/env python3
"""Load aggregate Barawell Growth OS history from an order-history XLSX export.

The script reads the flat `Data Mentah (per item)` tab, calculates deterministic
monthly/cohort/segment marts, and optionally upserts only aggregate rows to
Supabase. Phone numbers are used only for in-memory hashing and are never logged,
written to output, or persisted by this script.
"""

from __future__ import annotations

import argparse
import calendar
import hashlib
import json
import math
import os
import re
import sys
import urllib.error
import urllib.request
from collections import Counter, defaultdict
from dataclasses import dataclass, field
from datetime import date, datetime
from pathlib import Path
from typing import Any, Iterable

try:
    from openpyxl import load_workbook
except ImportError as exc:  # pragma: no cover - operational dependency check
    raise SystemExit("Install the pinned dependency: pip install openpyxl==3.1.5") from exc

MONTHS = {
    "jan": 1, "feb": 2, "mar": 3, "apr": 4, "mei": 5, "jun": 6,
    "jul": 7, "agu": 8, "sep": 9, "okt": 10, "nov": 11, "des": 12,
}
SUCCESS_STATUSES = {
    "delivered", "shipped", "paid", "completed", "complete", "selesai",
    "terkirim", "diproses",
}
PRODUCT_KEYS = {
    "s50": re.compile(r"sildenafil\s*50", re.I),
    "s100": re.compile(r"sildenafil\s*100", re.I),
    "baralast": re.compile(r"baralast", re.I),
    "tadalafil": re.compile(r"tadalafil", re.I),
}


@dataclass
class Order:
    customer_key: str
    order_id: str
    ordered_on: date
    name: str
    city: str
    channel: str
    revenue: float = 0
    boxes: float = 0
    products: set[str] = field(default_factory=set)


@dataclass
class Customer:
    key: str
    first_order: date
    last_order: date
    order_count: int
    lifetime_revenue: float
    total_boxes: float
    primary_channel: str
    products: set[str]


def parse_date(value: Any) -> date | None:
    if value is None or value == "":
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = str(value).strip()
    match = re.fullmatch(r"(\d{1,2})\s+([A-Za-z]+)\s+(\d{4})", text)
    if match:
        month_number = MONTHS.get(match.group(2).lower()[:3])
        if month_number:
            return date(int(match.group(3)), month_number, int(match.group(1)))
    try:
        return datetime.fromisoformat(text).date()
    except ValueError:
        return None


def number(value: Any) -> float:
    if value is None or value == "":
        return 0.0
    if isinstance(value, (int, float)):
        return float(value) if math.isfinite(float(value)) else 0.0
    cleaned = re.sub(r"[^0-9,.-]", "", str(value))
    if "." in cleaned and "," not in cleaned:
        parts = cleaned.split(".")
        cleaned = "".join(parts) if all(len(part) == 3 for part in parts[1:]) else cleaned
    if "," in cleaned and "." not in cleaned:
        parts = cleaned.split(",")
        cleaned = "".join(parts) if all(len(part) == 3 for part in parts[1:]) else cleaned.replace(",", ".")
    try:
        return float(cleaned)
    except ValueError:
        return 0.0


def normalize_phone(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float):
        return f"{value:.0f}"
    return re.sub(r"\D", "", str(value))


def customer_key(phone: Any, name: Any, city: Any) -> str:
    normalized_phone = normalize_phone(phone)
    source = normalized_phone or f"{str(name or '').strip().lower()}|{str(city or '').strip().lower()}"
    return hashlib.sha256(source.encode("utf-8")).hexdigest()


def normalize_channel(value: Any) -> str:
    channel = str(value or "other").strip().lower()
    aliases = {"web": "website", "wa": "whatsapp", "toko pedia": "tokopedia"}
    return aliases.get(channel, channel or "other")


def channel_group(channel: str) -> str:
    if channel in {"website", "whatsapp"}:
        return "owned"
    if channel in {"shopee", "tokopedia", "tiktok"}:
        return "marketplace"
    return "other"


def month_start(value: date) -> date:
    return date(value.year, value.month, 1)


def month_end(value: date) -> date:
    return date(value.year, value.month, calendar.monthrange(value.year, value.month)[1])


def add_month(value: date, months: int) -> date:
    total = value.year * 12 + value.month - 1 + months
    return date(total // 12, total % 12 + 1, 1)


def product_keys(text: str) -> set[str]:
    return {key for key, pattern in PRODUCT_KEYS.items() if pattern.search(text)}


def read_orders(path: Path) -> list[Order]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    if "Data Mentah (per item)" not in workbook.sheetnames:
        raise ValueError("Workbook is missing the 'Data Mentah (per item)' tab.")
    sheet = workbook["Data Mentah (per item)"]
    rows = sheet.iter_rows(values_only=True)
    headers = [str(value or "").strip() for value in next(rows)]
    index = {header: position for position, header in enumerate(headers)}
    required = {"Tanggal", "Order ID", "Nama", "HP", "Kota", "Channel", "Status", "Produk", "Qty (box)", "Total Order"}
    missing = sorted(required - set(index))
    if missing:
        raise ValueError(f"Missing required columns: {', '.join(missing)}")

    grouped: dict[tuple[str, str], Order] = {}
    for row in rows:
        status = str(row[index["Status"]] or "").strip().lower()
        if status not in SUCCESS_STATUSES:
            continue
        ordered_on = parse_date(row[index["Tanggal"]])
        order_id = str(row[index["Order ID"]] or "").strip()
        if not ordered_on or not order_id:
            continue
        name = str(row[index["Nama"]] or "").strip()
        city = str(row[index["Kota"]] or "-").strip() or "-"
        key = customer_key(row[index["HP"]], name, city)
        group_key = (key, order_id)
        current = grouped.get(group_key)
        if current is None:
            current = Order(
                customer_key=key,
                order_id=order_id,
                ordered_on=ordered_on,
                name=name,
                city=city,
                channel=normalize_channel(row[index["Channel"]]),
            )
            grouped[group_key] = current
        current.revenue = max(current.revenue, number(row[index["Total Order"]]))
        current.boxes += number(row[index["Qty (box)"]])
        current.products.update(product_keys(str(row[index["Produk"]] or "")))

    return sorted((order for order in grouped.values() if order.revenue > 0), key=lambda order: (order.ordered_on, order.order_id))


def build_customers(orders: list[Order]) -> dict[str, Customer]:
    by_customer: dict[str, list[Order]] = defaultdict(list)
    for order in orders:
        by_customer[order.customer_key].append(order)
    customers: dict[str, Customer] = {}
    for key, customer_orders in by_customer.items():
        customer_orders.sort(key=lambda order: (order.ordered_on, order.order_id))
        channels = Counter(order.channel for order in customer_orders)
        products = set().union(*(order.products for order in customer_orders))
        customers[key] = Customer(
            key=key,
            first_order=customer_orders[0].ordered_on,
            last_order=customer_orders[-1].ordered_on,
            order_count=len(customer_orders),
            lifetime_revenue=sum(order.revenue for order in customer_orders),
            total_boxes=sum(order.boxes for order in customer_orders),
            primary_channel=channels.most_common(1)[0][0],
            products=products,
        )
    return customers


def classify(customer: Customer, as_of: date) -> tuple[str, str, int]:
    recency = max(0, (as_of - customer.last_order).days)
    value_tier = "Whale" if customer.lifetime_revenue >= 2_000_000 else "Core" if customer.lifetime_revenue >= 400_000 else "Entry"
    if recency > 90:
        lifecycle = "Dormant"
    elif recency >= 46:
        lifecycle = "At-risk"
    elif customer.order_count >= 2:
        lifecycle = "Active-repeat"
    else:
        lifecycle = "New"
    return value_tier, lifecycle, recency


def aggregate(path: Path) -> dict[str, list[dict[str, Any]]]:
    orders = read_orders(path)
    if not orders:
        raise ValueError("No successful orders were found.")
    customers = build_customers(orders)
    minimum, maximum = orders[0].ordered_on, max(order.ordered_on for order in orders)
    source_batch = path.stem

    monthly: list[dict[str, Any]] = []
    current = month_start(minimum)
    while current <= month_start(maximum):
        month_orders = [order for order in orders if month_start(order.ordered_on) == current]
        if month_orders:
            keys = {order.customer_key for order in month_orders}
            new = [order for order in month_orders if month_start(customers[order.customer_key].first_order) == current]
            repeat = [order for order in month_orders if month_start(customers[order.customer_key].first_order) < current]
            new_keys = {order.customer_key for order in new}
            repeat_keys = {order.customer_key for order in repeat}
            owned = sum(1 for order in month_orders if channel_group(order.channel) == "owned")
            marketplace = sum(1 for order in month_orders if channel_group(order.channel) == "marketplace")
            revenue = sum(order.revenue for order in month_orders)
            monthly.append({
                "metric_month": current.isoformat(), "orders": len(month_orders), "revenue": revenue,
                "customers": len(keys), "aov": revenue / len(month_orders),
                "new_orders": len(new), "new_revenue": sum(order.revenue for order in new), "new_customers": len(new_keys),
                "repeat_orders": len(repeat), "repeat_revenue": sum(order.revenue for order in repeat), "repeat_customers": len(repeat_keys),
                "repeat_rate": len(repeat_keys) / len(keys),
                "repeat_revenue_share": sum(order.revenue for order in repeat) / revenue,
                "owned_orders": owned, "marketplace_orders": marketplace,
                "other_orders": len(month_orders) - owned - marketplace,
                "is_complete": month_end(current) <= maximum, "source_batch": source_batch,
            })
        current = add_month(current, 1)

    cohorts: list[dict[str, Any]] = []
    for cohort_month in sorted({month_start(customer.first_order) for customer in customers.values()}):
        cohort = [customer for customer in customers.values() if month_start(customer.first_order) == cohort_month]
        repeat_45 = repeat_90 = 0
        revenue_90 = 0.0
        for customer in cohort:
            customer_orders = sorted((order for order in orders if order.customer_key == customer.key), key=lambda order: order.ordered_on)
            first = customer_orders[0].ordered_on
            gaps = [(order.ordered_on - first).days for order in customer_orders[1:]]
            repeat_45 += int(any(gap <= 45 for gap in gaps))
            repeat_90 += int(any(gap <= 90 for gap in gaps))
            revenue_90 += sum(order.revenue for order in customer_orders if (order.ordered_on - first).days <= 90)
        count = len(cohort)
        cohorts.append({
            "cohort_month": cohort_month.isoformat(), "customers": count,
            "repeat_45": repeat_45, "repeat_90": repeat_90,
            "repeat_45_rate": repeat_45 / count, "repeat_90_rate": repeat_90 / count,
            "revenue_90": revenue_90, "revenue_90_per_customer": revenue_90 / count,
            "source_batch": source_batch,
        })

    channels: list[dict[str, Any]] = []
    products: list[dict[str, Any]] = []
    for metric in monthly:
        metric_month = date.fromisoformat(metric["metric_month"])
        month_orders = [order for order in orders if month_start(order.ordered_on) == metric_month]
        for channel in sorted({order.channel for order in month_orders}):
            channel_orders = [order for order in month_orders if order.channel == channel]
            channels.append({
                "metric_month": metric["metric_month"], "channel": channel, "channel_group": channel_group(channel),
                "orders": len(channel_orders), "customers": len({order.customer_key for order in channel_orders}),
                "revenue": sum(order.revenue for order in channel_orders), "source_batch": source_batch,
            })
        for product in PRODUCT_KEYS:
            product_orders = [order for order in month_orders if product in order.products]
            products.append({
                "metric_month": metric["metric_month"], "product": product,
                "orders": len(product_orders), "customers": len({order.customer_key for order in product_orders}),
                "revenue": sum(order.revenue for order in product_orders), "source_batch": source_batch,
            })

    segments: list[dict[str, Any]] = []
    snapshots: dict[date, dict[str, tuple[str, str]]] = {}
    current = month_start(minimum)
    while current <= month_start(maximum):
        as_of = min(month_end(current), maximum)
        available_orders = [order for order in orders if order.ordered_on <= as_of]
        available_customers = build_customers(available_orders)
        snapshot: dict[str, tuple[str, str]] = {}
        buckets: dict[tuple[str, str], list[Customer]] = defaultdict(list)
        for customer in available_customers.values():
            tier, lifecycle, _ = classify(customer, as_of)
            snapshot[customer.key] = (tier, lifecycle)
            buckets[(tier, lifecycle)].append(customer)
        snapshots[current] = snapshot
        for (tier, lifecycle), group in buckets.items():
            recencies = [(as_of - customer.last_order).days for customer in group]
            segments.append({
                "snapshot_month": current.isoformat(), "value_tier": tier, "lifecycle": lifecycle,
                "customers": len(group), "revenue": sum(customer.lifetime_revenue for customer in group),
                "avg_aov": sum(customer.lifetime_revenue / customer.order_count for customer in group) / len(group),
                "avg_recency": sum(recencies) / len(recencies), "source_batch": source_batch,
            })
        current = add_month(current, 1)

    movements: list[dict[str, Any]] = []
    snapshot_months = sorted(snapshots)
    for prior_month, current_month in zip(snapshot_months, snapshot_months[1:]):
        current_customers = build_customers([order for order in orders if order.ordered_on <= min(month_end(current_month), maximum)])
        buckets: dict[tuple[str, str, str, str], list[Customer]] = defaultdict(list)
        for key, prior_state in snapshots[prior_month].items():
            current_state = snapshots[current_month].get(key)
            if not current_state or key not in current_customers:
                continue
            buckets[(*prior_state, *current_state)].append(current_customers[key])
        for states, group in buckets.items():
            movements.append({
                "snapshot_month": current_month.isoformat(),
                "from_value_tier": states[0], "from_lifecycle": states[1],
                "to_value_tier": states[2], "to_lifecycle": states[3],
                "customers": len(group), "revenue": sum(customer.lifetime_revenue for customer in group),
                "source_batch": source_batch,
            })

    return {
        "growth_monthly_metrics": monthly,
        "growth_cohort_metrics": cohorts,
        "growth_channel_monthly": channels,
        "growth_product_monthly": products,
        "growth_segment_monthly": segments,
        "growth_segment_movement": movements,
    }


def chunks(rows: list[dict[str, Any]], size: int = 250) -> Iterable[list[dict[str, Any]]]:
    for index in range(0, len(rows), size):
        yield rows[index:index + size]


def post_json(url: str, service_key: str, payload: Any, prefer: str = "resolution=merge-duplicates") -> Any:
    request = urllib.request.Request(
        url,
        data=json.dumps(payload, separators=(",", ":")).encode("utf-8"),
        headers={
            "apikey": service_key,
            "Authorization": f"Bearer {service_key}",
            "Content-Type": "application/json",
            "Prefer": prefer,
        },
        method="POST",
    )
    try:
        with urllib.request.urlopen(request, timeout=60) as response:
            body = response.read().decode("utf-8")
            return json.loads(body) if body else None
    except urllib.error.HTTPError as exc:
        detail = exc.read().decode("utf-8", errors="replace")
        raise RuntimeError(f"Supabase request failed ({exc.code}): {detail}") from exc


def load_to_supabase(marts: dict[str, list[dict[str, Any]]]) -> None:
    base_url = (os.environ.get("SUPABASE_URL") or os.environ.get("NEXT_PUBLIC_SUPABASE_URL") or "").rstrip("/")
    service_key = os.environ.get("SUPABASE_SERVICE_ROLE_KEY", "")
    if not base_url or not service_key:
        raise RuntimeError("Set SUPABASE_URL or NEXT_PUBLIC_SUPABASE_URL and SUPABASE_SERVICE_ROLE_KEY.")
    conflicts = {
        "growth_monthly_metrics": "metric_month",
        "growth_cohort_metrics": "cohort_month",
        "growth_channel_monthly": "metric_month,channel",
        "growth_product_monthly": "metric_month,product",
        "growth_segment_monthly": "snapshot_month,value_tier,lifecycle",
        "growth_segment_movement": "snapshot_month,from_value_tier,from_lifecycle,to_value_tier,to_lifecycle",
    }
    for table, rows in marts.items():
        for batch in chunks(rows):
            endpoint = f"{base_url}/rest/v1/{table}?on_conflict={conflicts[table]}"
            post_json(endpoint, service_key, batch)
        print(f"Loaded {len(rows):,} aggregate rows into {table}.")
    post_json(f"{base_url}/rest/v1/rpc/refresh_growth_insights", service_key, {})


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--input", required=True, type=Path, help="Path to the Barawell XLSX export.")
    parser.add_argument("--load", action="store_true", help="Upsert aggregate marts to Supabase.")
    parser.add_argument("--dry-run", action="store_true", help="Calculate and validate without loading.")
    args = parser.parse_args()
    if not args.input.exists():
        parser.error(f"Input file does not exist: {args.input}")

    marts = aggregate(args.input)
    summary = {table: len(rows) for table, rows in marts.items()}
    print(json.dumps({"input": args.input.name, "aggregate_rows": summary}, indent=2))
    if args.load and not args.dry_run:
        load_to_supabase(marts)
    else:
        print("Dry run complete. No customer-level data or phone numbers were emitted.")
    return 0


if __name__ == "__main__":
    sys.exit(main())
